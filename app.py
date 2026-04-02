import streamlit as st
from openai import OpenAI
from dotenv import load_dotenv
import os
import json
import io
import base64
from datetime import datetime
from pathlib import Path

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

import csv

# Load environment variables
load_dotenv()

# Initialize OpenAI client
@st.cache_resource
def get_openai_client():
    return OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

client = get_openai_client()

# Page config
st.set_page_config(
    page_title="BIM-CHATBOT",
    page_icon="🔱",
    layout="wide"
)

# Load trident image as base64 for background (remove white bg, keep as clean silhouette)
def get_trident_b64():
    import base64 as b64lib
    from PIL import Image as PILImage
    import numpy as np
    trident_path = Path("photo/trident.jpg")
    if trident_path.exists():
        img = PILImage.open(trident_path).convert("RGBA")
        data = np.array(img)
        r, g, b = data[:,:,0], data[:,:,1], data[:,:,2]
        # Remove white/light background → fully transparent
        light_mask = (r > 160) & (g > 160) & (b > 160)
        data[light_mask] = [0, 0, 0, 0]
        # Make dark pixels (the trident) a soft gold for brand accent
        dark_mask = ~light_mask
        data[dark_mask, 0] = 240  # R
        data[dark_mask, 1] = 192  # G
        data[dark_mask, 2] = 0    # B
        data[dark_mask, 3] = 255
        img_out = PILImage.fromarray(data)
        buf = io.BytesIO()
        img_out.save(buf, format="PNG")
        return b64lib.b64encode(buf.getvalue()).decode("utf-8")
    return ""

trident_b64 = get_trident_b64()

# Professional clean design with Barbados colors
trident_bg = f'url("data:image/png;base64,{trident_b64}")' if trident_b64 else 'none'
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ===== BASE ===== */
[data-testid="stMain"] {{
    background: linear-gradient(135deg, #0A1F44, #133E7C) !important;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    position: relative;
}}

/* ===== TRIDENT BACKGROUND ===== */
[data-testid="stMain"]::before {{
    content: "";
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background-image: {trident_bg};
    background-repeat: no-repeat;
    background-position: center;
    background-size: 350px;
    opacity: 0.15;
    pointer-events: none;
    z-index: 0;
}}

/* ===== CONTENT LAYER ===== */
[data-testid="stMain"] > div {{
    position: relative;
    z-index: 1;
}}

/* ===== TEXT ===== */
[data-testid="stMain"] * {{
    color: #E6EDF3;
}}

/* ===== SIDEBAR ===== */
[data-testid="stSidebar"] {{
    background: #010409 !important;
    border-right: 1px solid #21262D !important;
}}
[data-testid="stSidebar"] * {{
    color: #C9D1D9 !important;
}}
[data-testid="stSidebar"] .stButton > button {{
    background: #161B22 !important;
    color: #F0C000 !important;
    border: 1px solid #30363D !important;
    border-radius: 8px;
    transition: all 0.2s ease;
    font-weight: 500;
    font-size: 14px;
    padding: 8px 16px;
}}
[data-testid="stSidebar"] .stButton > button:hover {{
    background: #1C2333 !important;
    border-color: #F0C000 !important;
    color: #F0C000 !important;
}}

/* ===== CHAT INPUT ===== */
[data-testid="stChatInput"] {{
    border-top: 1px solid #21262D !important;
}}
[data-testid="stChatInput"] textarea {{
    background: #161B22 !important;
    color: #E6EDF3 !important;
    border: 1px solid #30363D !important;
    border-radius: 12px;
    font-family: 'Inter', sans-serif;
    font-size: 15px;
    padding: 12px 16px;
}}
[data-testid="stChatInput"] textarea:focus {{
    border-color: #F0C000 !important;
    box-shadow: 0 0 0 1px rgba(240, 192, 0, 0.3) !important;
}}
[data-testid="stChatInput"] textarea::placeholder {{
    color: #484F58 !important;
}}

/* ===== CHAT MESSAGES ===== */
[data-testid="stChatMessage"] {{
    border-radius: 0 !important;
    padding: 20px 24px !important;
    margin-bottom: 0 !important;
    border: none !important;
    box-shadow: none !important;
    max-width: 100%;
}}
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]) {{
    background: #0D1117 !important;
    border-bottom: 1px solid #161B22 !important;
}}
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-assistant"]) {{
    background: #161B22 !important;
    border-bottom: 1px solid #21262D !important;
}}

/* Chat message text */
[data-testid="stChatMessage"] p,
[data-testid="stChatMessage"] li,
[data-testid="stChatMessage"] span {{
    font-size: 15px;
    line-height: 1.7;
    color: #E6EDF3 !important;
}}
[data-testid="stChatMessage"] code {{
    background: #0D1117 !important;
    color: #79C0FF !important;
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 13px;
}}
[data-testid="stChatMessage"] pre {{
    background: #0D1117 !important;
    border: 1px solid #21262D !important;
    border-radius: 8px;
    padding: 16px !important;
}}

/* ===== AVATAR STYLING ===== */
[data-testid="chatAvatarIcon-assistant"] {{
    background: #F0C000 !important;
    color: #0D1117 !important;
}}
[data-testid="chatAvatarIcon-user"] {{
    background: #1F6FEB !important;
}}

/* ===== TITLE ===== */
h1 {{
    color: #F0C000 !important;
    font-weight: 700;
    font-size: 28px !important;
    letter-spacing: 0.5px;
    text-shadow: none;
}}
h2 {{
    color: #F0C000 !important;
    font-weight: 600;
    font-size: 18px !important;
}}
h3 {{
    color: #C9D1D9 !important;
    font-weight: 600;
}}

/* ===== CAPTION ===== */
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] * {{
    color: #484F58 !important;
    font-size: 13px !important;
}}

/* ===== SELECTBOX ===== */
[data-testid="stSelectbox"] div[data-baseweb="select"] {{
    background: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 8px;
}}
[data-testid="stSelectbox"] div[data-baseweb="select"]:hover {{
    border-color: #484F58 !important;
}}

/* ===== TEXT INPUT ===== */
[data-testid="stTextInput"] input {{
    background: #161B22 !important;
    color: #E6EDF3 !important;
    border: 1px solid #30363D !important;
    border-radius: 8px;
    font-size: 14px;
}}
[data-testid="stTextInput"] input:focus {{
    border-color: #F0C000 !important;
    box-shadow: 0 0 0 1px rgba(240, 192, 0, 0.3) !important;
}}

/* ===== FILE UPLOADER ===== */
[data-testid="stFileUploader"] {{
    border: 1px dashed #30363D !important;
    border-radius: 8px;
    padding: 12px;
    transition: border-color 0.2s ease;
    background: #161B22 !important;
}}
[data-testid="stFileUploader"]:hover {{
    border-color: #F0C000 !important;
}}

/* ===== DIVIDERS ===== */
hr {{
    border-color: #21262D !important;
}}

/* ===== SCROLLBAR ===== */
::-webkit-scrollbar {{ width: 8px; }}
::-webkit-scrollbar-track {{ background: #0D1117; }}
::-webkit-scrollbar-thumb {{ background: #30363D; border-radius: 4px; }}
::-webkit-scrollbar-thumb:hover {{ background: #484F58; }}

/* ===== LINKS ===== */
a {{ color: #58A6FF !important; text-decoration: none !important; }}
a:hover {{ color: #79C0FF !important; text-decoration: underline !important; }}

/* ===== ALERTS ===== */
[data-testid="stAlert"] {{
    border-radius: 8px;
    background: #161B22 !important;
    border: 1px solid #30363D !important;
}}

/* ===== STREAMLIT HEADER HIDE ===== */
header[data-testid="stHeader"] {{
    background: #0D1117 !important;
    border-bottom: 1px solid #21262D;
}}

/* ===== BOTTOM BAR ===== */
[data-testid="stBottomBlockContainer"] {{
    background: #0D1117 !important;
}}

/* ===== DROPDOWN MENU ===== */
[data-baseweb="menu"],
[data-baseweb="popover"] {{
    background: #161B22 !important;
    border: 1px solid #30363D !important;
}}
[data-baseweb="menu"] li {{
    color: #E6EDF3 !important;
}}
[data-baseweb="menu"] li:hover {{
    background: #21262D !important;
}}

/* ===== SIDEBAR LABELS ===== */
[data-testid="stSidebar"] label {{
    color: #8B949E !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}}

/* ===== WRITE TEXT ===== */
[data-testid="stSidebar"] [data-testid="stText"] {{
    color: #8B949E !important;
    font-size: 13px !important;
}}
</style>
""", unsafe_allow_html=True)

# Load upcoming events
EVENTS_FILE = Path("events.json")

def load_events():
    if EVENTS_FILE.exists():
        with open(EVENTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

ALL_EVENTS = load_events()

# Chat history file management
HISTORY_DIR = Path("chat_histories")
HISTORY_DIR.mkdir(exist_ok=True)

def save_conversation(name, messages):
    """Save a conversation to a JSON file."""
    filepath = HISTORY_DIR / f"{name}.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump({"name": name, "messages": messages, "saved_at": datetime.now().isoformat()}, f, indent=2)

def load_conversation(name):
    """Load a conversation from a JSON file."""
    filepath = HISTORY_DIR / f"{name}.json"
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["messages"]

def list_conversations():
    """List all saved conversations."""
    convos = []
    for f in sorted(HISTORY_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        with open(f, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        convos.append({"name": f.stem, "saved_at": data.get("saved_at", "")})
    return convos

def delete_conversation(name):
    """Delete a saved conversation."""
    filepath = HISTORY_DIR / f"{name}.json"
    filepath.unlink(missing_ok=True)

# Initialize session state
if "messages" not in st.session_state:
    st.session_state.messages = []
if "documents" not in st.session_state:
    st.session_state.documents = []
if "current_chat_name" not in st.session_state:
    st.session_state.current_chat_name = None
if "images" not in st.session_state:
    st.session_state.images = []
if "selected_model" not in st.session_state:
    st.session_state.selected_model = "gpt-4o"

# Available models
MODELS = {
    "GPT-4o (Best accuracy)": "gpt-4o",
    "GPT-4o Mini (Fast & cheap)": "gpt-4o-mini",
    "GPT-4 Turbo": "gpt-4-turbo",
    "GPT-4": "gpt-4",
    "GPT-3.5 Turbo (Budget)": "gpt-3.5-turbo",
    "o1 (Advanced reasoning)": "o1",
    "o1-mini (Fast reasoning)": "o1-mini",
    "o3-mini (Latest reasoning)": "o3-mini",
}

# Sidebar
with st.sidebar:
    # --- New Chat ---
    st.title("🏗️ BIM-CHATBOT")

    # --- Model Selector ---
    selected_label = st.selectbox(
        "🧠 AI Model",
        options=list(MODELS.keys()),
        index=list(MODELS.values()).index(st.session_state.selected_model),
        help="Choose the AI model. Better models cost more but are more accurate."
    )
    st.session_state.selected_model = MODELS[selected_label]

    if st.button("➕ New Chat", use_container_width=True):
        # Auto-save current chat before starting new one
        if st.session_state.messages:
            auto_name = datetime.now().strftime("Chat %Y-%m-%d %H:%M")
            if st.session_state.current_chat_name:
                auto_name = st.session_state.current_chat_name
            save_conversation(auto_name, st.session_state.messages)
        st.session_state.messages = []
        st.session_state.current_chat_name = None
        st.rerun()

    # --- Chat History ---
    st.divider()
    st.subheader("💬 Chat History")
    conversations = list_conversations()
    if conversations:
        for convo in conversations:
            col1, col2 = st.columns([4, 1])
            with col1:
                if st.button(f"📄 {convo['name']}", key=f"load_{convo['name']}", use_container_width=True):
                    st.session_state.messages = load_conversation(convo["name"])
                    st.session_state.current_chat_name = convo["name"]
                    st.rerun()
            with col2:
                if st.button("🗑️", key=f"del_{convo['name']}"):
                    delete_conversation(convo["name"])
                    if st.session_state.current_chat_name == convo["name"]:
                        st.session_state.current_chat_name = None
                    st.rerun()
    else:
        st.caption("No saved chats yet.")

    # --- Save Current Chat ---
    st.divider()
    save_name = st.text_input("Save chat as:", value=st.session_state.current_chat_name or "")
    if st.button("💾 Save Chat", use_container_width=True):
        if save_name and st.session_state.messages:
            save_conversation(save_name, st.session_state.messages)
            st.session_state.current_chat_name = save_name
            st.success(f"✓ Saved: {save_name}")
            st.rerun()

    # --- Document Upload ---
    st.divider()
    st.subheader("📁 Document Upload")
    uploaded_file = st.file_uploader(
        "Upload any file",
        type=['txt', 'md', 'pdf', 'docx', 'doc', 'csv', 'xlsx', 'xls', 'json', 'xml', 'html', 'py', 'js', 'ts', 'java', 'c', 'cpp', 'h', 'css', 'sql', 'yaml', 'yml', 'toml', 'ini', 'cfg', 'log', 'rtf', 'jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp', 'svg'],
        help="Upload documents or images for the AI to reference"
    )
    
    if uploaded_file is not None:
        content = None
        fname = uploaded_file.name.lower()
        
        try:
            if fname.endswith('.pdf'):
                if PyPDF2:
                    reader = PyPDF2.PdfReader(uploaded_file)
                    content = "".join(page.extract_text() or "" for page in reader.pages)
                else:
                    st.warning("Install PyPDF2 to read PDFs: pip install PyPDF2")
            elif fname.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
                img_bytes = uploaded_file.read()
                b64 = base64.b64encode(img_bytes).decode('utf-8')
                ext = fname.rsplit('.', 1)[-1]
                mime = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"
                img_data = {"name": uploaded_file.name, "b64": b64, "mime": mime}
                if img_data not in st.session_state.images:
                    st.session_state.images.append(img_data)
                    st.success(f"✓ Image added: {uploaded_file.name}")
                    st.image(img_bytes, caption=uploaded_file.name, width=200)
                content = None  # Images handled separately
            elif fname.endswith('.svg'):
                content = uploaded_file.read().decode('utf-8', errors='replace')
            elif fname.endswith('.docx'):
                if docx:
                    doc_file = docx.Document(uploaded_file)
                    content = "\n".join(p.text for p in doc_file.paragraphs)
                else:
                    st.warning("Install python-docx to read Word files: pip install python-docx")
            elif fname.endswith(('.xlsx', '.xls')):
                if openpyxl:
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheets_text = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        sheets_text.append(f"--- Sheet: {sheet_name} ---")
                        for row in ws.iter_rows(values_only=True):
                            sheets_text.append("\t".join(str(cell) if cell is not None else "" for cell in row))
                    content = "\n".join(sheets_text)
                else:
                    st.warning("Install openpyxl to read Excel files: pip install openpyxl")
            elif fname.endswith('.csv'):
                text = uploaded_file.read().decode('utf-8', errors='replace')
                content = text
            elif fname.endswith('.json'):
                raw = uploaded_file.read().decode('utf-8', errors='replace')
                data = json.loads(raw)
                content = json.dumps(data, indent=2)
            else:
                # Try to read as plain text
                content = uploaded_file.read().decode('utf-8', errors='replace')
        except Exception as e:
            st.error(f"Error reading {uploaded_file.name}: {e}")
        
        if content and content.strip():
            if content not in st.session_state.documents:
                st.session_state.documents.append(content)
                st.success(f"✓ Added: {uploaded_file.name}")
        elif content is not None and not content.strip():
            st.warning(f"No readable text found in {uploaded_file.name}")
    
    st.write(f"**Documents loaded:** {len(st.session_state.documents)}")
    st.write(f"**Images loaded:** {len(st.session_state.images)}")
    
    if st.button("🗑️ Clear Documents"):
        st.session_state.documents = []
        st.session_state.images = []
        st.rerun()

# Main chat interface
st.title("🏗️ BIM-CHATBOT")
st.caption("Your intelligent BIM assistant powered by document retrieval")

# Upcoming events panel
if ALL_EVENTS:
    with st.expander("🎉 Upcoming Barbados Events", expanded=False):
        # Sort: weekly events last, then by date string
        upcoming = [e for e in ALL_EVENTS if e.get("category") != "Weekly"]
        weekly = [e for e in ALL_EVENTS if e.get("category") == "Weekly"]
        cols = st.columns(min(len(upcoming[:4]), 4)) if upcoming else []
        for i, event in enumerate(upcoming[:4]):
            with cols[i % 4]:
                st.markdown(f"**{event['name']}**")
                st.caption(f"📅 {event['date']}")
                st.caption(f"📍 {event['location']}")
                st.caption(f"🎟️ {event['cost']}")
                if event.get('link'):
                    st.markdown(f"[More info]({event['link']})")
        if len(upcoming) > 4:
            st.divider()
            for event in upcoming[4:]:
                st.markdown(f"**{event['name']}** — {event['date']} | 📍 {event['location']} | 🎟️ {event['cost']}")
        if weekly:
            st.divider()
            st.markdown("**🔁 Weekly Events**")
            for event in weekly:
                st.markdown(f"**{event['name']}** — {event['date']} | 📍 {event['location']} | 🎟️ {event['cost']}")

# Display chat messages
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

# Retrieve relevant context from documents
def retrieve_context(query: str, top_k: int = 2) -> str:
    if not st.session_state.documents:
        return ""
    
    query_words = set(query.lower().split())
    scored_docs = []
    
    for doc in st.session_state.documents:
        doc_words = set(doc.lower().split())
        score = len(query_words & doc_words)
        scored_docs.append((score, doc))
    
    scored_docs.sort(reverse=True, key=lambda x: x[0])
    relevant = [doc for score, doc in scored_docs[:top_k] if score > 0]
    
    return "\n\n".join(relevant) if relevant else ""

# Chat input
if prompt := st.chat_input("Ask me anything..."):
    # Retrieve context from documents
    context = retrieve_context(prompt)
    
    # Build message with context
    if context:
        enhanced_prompt = f"Context from documents:\n{context}\n\nUser question: {prompt}"
        display_prompt = prompt  # Show original to user
    else:
        enhanced_prompt = prompt
        display_prompt = prompt
    
    # Add user message to chat
    st.session_state.messages.append({"role": "user", "content": enhanced_prompt})
    
    # Display user message
    with st.chat_message("user"):
        st.markdown(display_prompt)
    
    # Get AI response
    with st.chat_message("assistant"):
        message_placeholder = st.empty()
        
        try:
            # Build messages with system prompt including current date/time
            now = datetime.now()
            system_msg = {
                "role": "system",
                "content": (
                    f"You are BIM-CHATBOT, an expert-level AI assistant that prioritizes ACCURACY above all else. "
                    f"The current date and time is {now.strftime('%A, %B %d, %Y at %I:%M %p')}.\n\n"
                    f"CORE RULES:\n"
                    f"- ACCURACY IS YOUR #1 PRIORITY. Never guess or make up information.\n"
                    f"- If you are not 100% sure about something, say so honestly.\n"
                    f"- Always think through your answer carefully before responding.\n\n"
                    f"MATH INSTRUCTIONS:\n"
                    f"- ALWAYS solve math problems step by step with clear labels (Step 1, Step 2, etc.).\n"
                    f"- Show ALL work — do not skip steps.\n"
                    f"- Double-check every calculation before giving your final answer.\n"
                    f"- State the FINAL ANSWER in bold on its own line.\n"
                    f"- Explain WHY each step works so the user learns.\n"
                    f"- For word problems, identify what is given, what is asked, then solve.\n"
                    f"- Use proper mathematical notation and formatting.\n\n"
                    f"ENGLISH / GRAMMAR / WRITING:\n"
                    f"- Provide grammatically perfect responses.\n"
                    f"- When correcting grammar, explain the rule behind the correction.\n"
                    f"- For essays or writing help, follow proper structure (intro, body, conclusion).\n"
                    f"- When asked about vocabulary, give clear definitions, synonyms, and example sentences.\n"
                    f"- For literature questions, cite specific evidence from the text.\n\n"
                    f"SCIENCE / HISTORY / GENERAL KNOWLEDGE:\n"
                    f"- Provide factual, well-sourced information.\n"
                    f"- Use specific dates, names, and data — not vague statements.\n"
                    f"- Distinguish between established facts and theories/opinions.\n"
                    f"- If a topic is debated, present multiple perspectives fairly.\n\n"
                    f"FORMATTING:\n"
                    f"- Use headers, bullet points, and bold text for readability.\n"
                    f"- Keep explanations clear and well-organized.\n"
                    f"- Adapt your explanation depth to the complexity of the question.\n\n"
                    f"BARBADOS EXPERT KNOWLEDGE:\n"
                    f"You are also a Barbados travel, culture, and lifestyle expert. When anyone asks about Barbados, provide rich, detailed, and helpful answers.\n\n"
                    f"HISTORY OF BARBADOS:\n"
                    f"- Barbados was inhabited by Arawak and Carib peoples before European contact.\n"
                    f"- The Portuguese visited in the 1500s and named it 'Los Barbados' (the bearded ones) after the fig trees.\n"
                    f"- British colonized Barbados in 1627. It became a major sugar colony using enslaved African labor.\n"
                    f"- Bussa's Rebellion (1816) was a major slave uprising. Emancipation came in 1834.\n"
                    f"- Barbados gained independence on November 30, 1966 (Independence Day is a national holiday).\n"
                    f"- Barbados became a Republic on November 30, 2021, removing the British monarch as head of state.\n"
                    f"- Dame Sandra Mason became the first President. The Prime Minister is Mia Amor Mottley.\n"
                    f"- National heroes include Errol Barrow, Sir Garfield Sobers, Sir Frank Walcott, Bussa, Sarah Ann Gill, Samuel Jackman Prescod, Charles Duncan O'Neal, Clement Payne, Sir Hugh Springer, and Rihanna.\n\n"
                    f"BEACHES & PLACES TO VISIT:\n"
                    f"- When asked about beaches, provide: Crane Beach, Bathsheba Beach, Bottom Bay, Miami Beach (Enterprise), Accra Beach (Rockley), Brownes Beach, Mullins Beach, Paynes Bay, Carlisle Bay, Silver Sands.\n"
                    f"- Popular attractions: Harrison's Cave, Animal Flower Cave, Hunte's Gardens, Andromeda Botanic Gardens, St. Nicholas Abbey, George Washington House, Barbados Wildlife Reserve, Welchman Hall Gully, Farley Hill National Park, the Barbados Museum.\n"
                    f"- Historic sites: Garrison Historic Area (UNESCO World Heritage), Parliament Buildings, Bridgetown's Cheapside Market, Holetown (first settlement), Oistins Fish Fry.\n\n"
                    f"HOTELS & ACCOMMODATION (always provide real links):\n"
                    f"- Luxury: Sandy Lane Hotel (https://www.sandylane.com), The Crane Resort (https://www.thecrane.com), Cobblers Cove (https://www.cobblerscove.com), Colony Club (https://www.eleganthotels.com/colony-club)\n"
                    f"- Mid-range: Courtyard by Marriott (https://www.marriott.com), Hilton Barbados (https://www.hilton.com), Radisson Aquatica (https://www.radissonhotels.com), Accra Beach Hotel\n"
                    f"- Budget: Island Inn Hotel, Yellow Bird Hotel, Worthing Court Apartment Hotel\n"
                    f"- Booking sites: https://www.visitbarbados.org, https://www.booking.com, https://www.airbnb.com\n\n"
                    f"VILLAS & RENTALS:\n"
                    f"- Villa rental sites: https://www.barbadosdreamvillas.com, https://www.vrbo.com, https://www.airbnb.com, https://www.terracaribbean.com/barbados/rentals, https://www.realtorslimited.com\n"
                    f"- Areas for villas: Royal Westmoreland, Sugar Hill, Sandy Lane Estate, Apes Hill, Port St. Charles, St. James coast, Christ Church south coast.\n"
                    f"- Always mention price ranges if possible and recommend checking the sites for latest availability.\n\n"
                    f"LAND FOR SALE:\n"
                    f"- Real estate sites: https://www.terracaribbean.com, https://www.realtorslimited.com, https://www.caribbeanluxuryproperty.com, https://www.propertiesinbarbados.com\n"
                    f"- Popular areas: West Coast (St. James, St. Peter), South Coast (Christ Church), Apes Hill, Royal Westmoreland.\n\n"
                    f"PLACES TO RENT:\n"
                    f"- Rental listings: https://www.terracaribbean.com/barbados/rentals, https://www.realtorslimited.com, https://www.caribbeanluxuryproperty.com, Facebook Marketplace Barbados, https://www.airbnb.com (long-term)\n"
                    f"- Recommend contacting local real estate agents for best deals.\n\n"
                    f"RENTAL CARS:\n"
                    f"- Companies: Stoutes Car Rental (https://www.stoutescar.com), Direct Car Rentals (https://www.directcarbarbados.com), Drive-A-Matic (https://www.driveamatic.com), Courtesy Rent-A-Car (https://www.courtesyrentacar.com)\n"
                    f"- Note: Driving is on the LEFT side of the road. You need a Barbados driving permit (available at rental agencies or police stations for about $10 BBD).\n\n"
                    f"JOBS IN BARBADOS:\n"
                    f"- Job sites: https://www.caribbeanjobs.com, https://www.glassdoor.com, https://www.linkedin.com/jobs (search Barbados), https://www.indeed.com\n"
                    f"- Government jobs: https://www.barbadosparliament.com, check the Barbados Government Information Service\n"
                    f"- Top industries: Tourism & hospitality, financial services, technology, construction, education, healthcare.\n"
                    f"- Mention the Barbados Welcome Stamp for remote workers (https://www.barbadoswelcomestamp.bb).\n\n"
                    f"SCHOOLS IN BARBADOS:\n"
                    f"- Top secondary schools: Harrison College, Queen's College, Combermere School, The Lodge School, Christ Church Foundation, Coleridge & Parry School, Alexandra School.\n"
                    f"- Primary schools: Government primary schools island-wide, private options like Codrington School.\n"
                    f"- University: University of the West Indies Cave Hill Campus (https://www.cavehill.uwi.edu), Barbados Community College (https://www.bcc.edu.bb)\n"
                    f"- International schools: The Codrington School (https://www.codrington.edu.bb)\n\n"
                    f"BAJAN COOKING:\n"
                    f"- When asked about Bajan cooking, provide full recipes with ingredients and steps.\n"
                    f"- Signature dishes: Cou-cou & Flying Fish (national dish), Macaroni Pie, Rice & Peas, Fish Cakes, Pudding & Souse, Conkies, Bajan Pepper Sauce, Bajan Seasoning, Breadfruit cou-cou, Jug Jug, Black Cake (Christmas).\n"
                    f"- Drinks: Rum Punch (Bajan style), Mauby, Sorrel, Coconut Water, Banks Beer.\n"
                    f"- Key seasonings: Bajan seasoning (thyme, marjoram, green onion, scotch bonnet, garlic, lime juice).\n"
                    f"- Always explain the cultural significance of the dish.\n\n"
                    f"TOUR GUIDES & ADVENTURES:\n"
                    f"- Island Safari (https://www.islandsafari.bb) - Jeep tours\n"
                    f"- Atlantis Submarines (https://www.atlantissubmarines.com) - underwater tours\n"
                    f"- Cool Runnings Catamaran Cruises - snorkeling with turtles\n"
                    f"- Harbour Master Cruises - party cruises\n"
                    f"- Adventures: Zip-lining, surfing at Bathsheba, snorkeling at Carlisle Bay (shipwrecks & turtles), hiking at Welchman Hall Gully, horseback riding, jet skiing, paddleboarding.\n"
                    f"- Harrison's Cave tours: https://www.harrisonscave.com\n\n"
                    f"CHURCHES:\n"
                    f"- St. John's Parish Church (stunning cliffside views), St. Michael's Cathedral, St. James Parish Church (oldest church, built 1628), St. George Parish Church, Bridgetown Synagogue (one of oldest in the Western Hemisphere).\n"
                    f"- Note Barbados has a strong Christian heritage with Anglican, Catholic, Methodist, Pentecostal, and other denominations.\n\n"
                    f"BARS & NIGHTLIFE:\n"
                    f"- Oistins Fish Fry (Friday night is legendary), St. Lawrence Gap (bar strip on south coast), Holetown nightlife.\n"
                    f"- Popular bars: The Boatyard, Harbour Lights, Naru Restaurant & Lounge, Tapas, Red Door Lounge, Nikki Beach.\n"
                    f"- Rum shops are a Barbados cultural institution — small local bars found everywhere.\n"
                    f"- Mount Gay Rum Distillery tours (https://www.mountgayrum.com) — oldest rum company in the world (since 1703).\n\n"
                    f"GAMES & SPORTS:\n"
                    f"- Cricket is the national sport. Kensington Oval is the main cricket ground.\n"
                    f"- Other popular sports: football (soccer), road tennis (invented in Barbados!), horse racing at the Garrison Savannah, surfing, basketball.\n"
                    f"- Road tennis is unique to Barbados — played on the road with wooden paddles.\n\n"
                    f"UPCOMING BARBADOS EVENTS (use this data when users ask about events, things to do, festivals, etc.):\n"
                    f"{chr(10).join(f'- {e["name"]}: {e["date"]} | Location: {e["location"]} | Cost: {e["cost"]} | {e["description"]}' + (f' | Link: {e["link"]}' if e.get("link") else '') for e in ALL_EVENTS)}\n\n"
                    f"IMPORTANT LINKS RULE:\n"
                    f"- When providing recommendations, ALWAYS include real, working website links where available.\n"
                    f"- When asked about events, festivals, or things to do, ALWAYS reference the upcoming events list above with dates, locations, ticket costs, and links.\n"
                    f"- Provide reviews and ratings when you can (e.g., 'highly rated', 'popular with visitors').\n"
                    f"- Recommend checking https://www.visitbarbados.org for the latest official tourism info.\n"
                    f"- If you don't have a specific link, say so and suggest where to search."
                )
            }
            api_messages = [system_msg] + st.session_state.messages

            # If images are uploaded, attach them to the last user message for vision
            if st.session_state.images:
                # Build multimodal content for the last user message
                last_user_idx = None
                for i in range(len(api_messages) - 1, -1, -1):
                    if api_messages[i]["role"] == "user":
                        last_user_idx = i
                        break
                if last_user_idx is not None:
                    text_content = api_messages[last_user_idx]["content"]
                    multimodal_content = [{"type": "text", "text": text_content}]
                    for img in st.session_state.images:
                        multimodal_content.append({
                            "type": "image_url",
                            "image_url": {"url": f"data:{img['mime']};base64,{img['b64']}"}
                        })
                    api_messages[last_user_idx] = {"role": "user", "content": multimodal_content}

            # Stream response for ChatGPT-like effect
            use_model = st.session_state.selected_model
            # o1/o3 models don't support streaming or temperature
            is_reasoning = use_model.startswith(("o1", "o3"))

            if is_reasoning:
                response = client.chat.completions.create(
                    model=use_model,
                    messages=api_messages,
                    max_completion_tokens=4000,
                )
                full_response = response.choices[0].message.content
                message_placeholder.markdown(full_response)
            else:
                stream = client.chat.completions.create(
                    model=use_model,
                    messages=api_messages,
                    temperature=0.1,
                    max_tokens=4000,
                    stream=True
                )
                
                full_response = ""
                for chunk in stream:
                    if chunk.choices[0].delta.content is not None:
                        full_response += chunk.choices[0].delta.content
                        message_placeholder.markdown(full_response + "▌")
                
                message_placeholder.markdown(full_response)
            
            # Add assistant response to history
            st.session_state.messages.append({"role": "assistant", "content": full_response})

            # Auto-save after each response
            if st.session_state.current_chat_name:
                save_conversation(st.session_state.current_chat_name, st.session_state.messages)
            
        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.info("💡 Make sure your API key is valid in the .env file")
